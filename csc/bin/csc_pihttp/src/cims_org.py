"""
CIMS 조직 관리 REST API
GET/POST/PUT/DELETE /api/v1/organizations
POST /api/v1/organizations/import (Excel)
GET  /api/v1/organizations/import/template
DELETE /api/v1/organizations/batch
"""

import json
import io
from urllib.parse import urlparse, parse_qs, unquote
from pathlib import PurePath

import pymysql
import pymysql.cursors

from util.pi_http.http_handler import HandlerArgs, HandlerResult

_ORG_BASE = '/api/v1/organizations'


def _get_db(config: dict):
    db = config.get('CimsDatabase', {})
    return pymysql.connect(
        host=db.get('Host', '127.0.0.1'),
        port=int(db.get('Port', 3306)),
        user=db.get('User', 'root'),
        password=db.get('Password', ''),
        database=db.get('Db', 'cims'),
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
    )


def _path_parts(full_path: str, base: str):
    path = urlparse(full_path).path
    try:
        rel = PurePath(path).relative_to(PurePath(base))
        return tuple(unquote(p) for p in rel.parts)
    except ValueError:
        return ()


async def handle_organizations(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    config = kwargs.get('config', {})
    parts = _path_parts(handler_args.full_path, _ORG_BASE)
    method = handler_args.method.upper()

    try:
        if len(parts) == 0:
            if method == 'GET':
                return await _list_orgs(config)
            elif method == 'POST':
                return await _create_org(handler_args.body, config)
            return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        if parts[0] == 'batch' and method == 'DELETE':
            return await _batch_delete_orgs(handler_args.body, config)

        if parts[0] == 'import':
            if len(parts) >= 2 and parts[1] == 'template' and method == 'GET':
                return _generate_template()
            if method == 'POST':
                return await _import_orgs(handler_args, config)
            return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        org_id = parts[0]
        if len(parts) == 1:
            if method == 'GET':
                return await _get_org(org_id, config)
            elif method == 'PUT':
                return await _update_org(org_id, handler_args.body, config)
            elif method == 'DELETE':
                return await _delete_org(org_id, config)

        if len(parts) == 2 and parts[1] == 'users' and method == 'GET':
            return await _list_org_users(org_id, config)

        return HandlerResult(status=404, body={'error': 'Not Found'})
    except pymysql.Error as e:
        return HandlerResult(status=500, body={'error': str(e)})


async def _list_orgs(config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, code, code_path, name, parent_id, sort_order FROM organizations ORDER BY code_path, sort_order, name"
            )
            rows = cur.fetchall()
    return HandlerResult(status=200, body={'organizations': rows})


async def _get_org(org_id, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM organizations WHERE id=%s", (org_id,))
            row = cur.fetchone()
            if not row:
                return HandlerResult(status=404, body={'error': 'Not found'})
    return HandlerResult(status=200, body=row)


def _build_code_path(cur, parent_id, code):
    """parent_id로부터 code_path 계산"""
    if not parent_id:
        return code
    cur.execute("SELECT code_path FROM organizations WHERE id=%s", (parent_id,))
    row = cur.fetchone()
    parent_path = row['code_path'] if row and row.get('code_path') else ''
    return (parent_path + '/' + code) if parent_path else code


def _rebuild_children_paths(cur, parent_id, parent_path):
    """부모 변경 시 모든 하위 조직의 code_path 재계산"""
    cur.execute("SELECT id, code FROM organizations WHERE parent_id=%s", (parent_id,))
    for child in cur.fetchall():
        child_path = parent_path + '/' + child['code']
        cur.execute("UPDATE organizations SET code_path=%s WHERE id=%s", (child_path, child['id']))
        _rebuild_children_paths(cur, child['id'], child_path)


async def _create_org(body, config):
    if not body:
        return HandlerResult(status=400, body={'error': 'Body required'})
    code = body.get('code', '').strip()
    name = body.get('name', '').strip()
    parent_id = body.get('parent_id')
    sort_order = body.get('sort_order', 0)
    if not code or not name:
        return HandlerResult(status=400, body={'error': 'code, name 필수'})

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            code_path = _build_code_path(cur, parent_id, code)
            cur.execute(
                "INSERT INTO organizations (code, code_path, name, parent_id, sort_order) "
                "VALUES (%s,%s,%s,%s,%s)",
                (code, code_path, name, parent_id, sort_order)
            )
            new_id = cur.lastrowid
    return HandlerResult(status=201, body={'id': new_id, 'code': code, 'code_path': code_path})


async def _update_org(org_id, body, config):
    if not body:
        return HandlerResult(status=400, body={'error': 'Body required'})
    sets, params = [], []
    for field in ('code', 'name', 'parent_id', 'sort_order'):
        if field in body:
            sets.append(f"{field}=%s")
            params.append(body[field])
    if not sets:
        return HandlerResult(status=400, body={'error': '변경할 필드가 없습니다'})
    params.append(org_id)
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE organizations SET {','.join(sets)} WHERE id=%s", params)
            # code 또는 parent_id 변경 시 code_path 재계산
            if 'parent_id' in body or 'code' in body:
                cur.execute("SELECT code, parent_id FROM organizations WHERE id=%s", (org_id,))
                row = cur.fetchone()
                if row:
                    new_path = _build_code_path(cur, row['parent_id'], row['code'])
                    cur.execute("UPDATE organizations SET code_path=%s WHERE id=%s", (new_path, org_id))
                    _rebuild_children_paths(cur, int(org_id), new_path)
    return HandlerResult(status=200, body={'id': org_id})


async def _delete_org(org_id, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            # 하위 조직의 parent를 NULL로 변경 (상위로 이동)
            cur.execute("UPDATE organizations SET parent_id=NULL WHERE parent_id=%s", (org_id,))
            cur.execute("DELETE FROM organizations WHERE id=%s", (org_id,))
            if cur.rowcount == 0:
                return HandlerResult(status=404, body={'error': 'Not found'})
    return HandlerResult(status=200, body={'id': org_id})


async def _batch_delete_orgs(body, config):
    ids = body.get('ids', []) if body else []
    if not ids:
        return HandlerResult(status=400, body={'error': 'ids 필요'})
    deleted = 0
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            for oid in ids:
                cur.execute("UPDATE organizations SET parent_id=NULL WHERE parent_id=%s", (oid,))
                cur.execute("DELETE FROM organizations WHERE id=%s", (oid,))
                if cur.rowcount > 0:
                    deleted += 1
    return HandlerResult(status=200, body={'deleted': deleted})


async def _list_org_users(org_id, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            # org_id로 code 조회
            cur.execute("SELECT code FROM organizations WHERE id=%s", (org_id,))
            row = cur.fetchone()
            if not row:
                return HandlerResult(status=404, body={'error': 'Organization not found'})
            code = row['code']
            cur.execute(
                "SELECT id, name, login_id, org_id FROM users WHERE org_id=%s ORDER BY name",
                (code,)
            )
            users = cur.fetchall()
    return HandlerResult(status=200, body={'org_id': org_id, 'org_code': code, 'users': users})


def _generate_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'organizations'
    ws.append(['code', 'name', 'parent_code', 'sort_order'])
    ws.append(['HQ', '본부', '', 1])
    ws.append(['DEV', '개발부', 'HQ', 1])
    ws.append(['DEV_01', '개발1팀', 'DEV', 1])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return HandlerResult(status=200, body=buf.getvalue(), headers={
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename="cims_org_template.xlsx"',
    })


async def _import_orgs(handler_args, config):
    import openpyxl
    import base64

    body = handler_args.body or {}
    file_data = None
    if 'file_base64' in body:
        file_data = base64.b64decode(body['file_base64'])
    if not file_data:
        return HandlerResult(status=400, body={'error': 'file_base64 필요'})

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data))
    except Exception as e:
        return HandlerResult(status=400, body={'error': f'파싱 실패: {e}'})

    if 'organizations' not in wb.sheetnames:
        return HandlerResult(status=400, body={'error': 'organizations 시트 필요'})

    ws = wb['organizations']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rd = dict(zip(headers, row))
        code = str(rd.get('code', '') or '').strip()
        name = str(rd.get('name', '') or '').strip()
        parent_code = str(rd.get('parent_code', '') or '').strip()
        sort_order = int(rd.get('sort_order', 0) or 0)
        if not code or not name:
            continue
        rows_data.append({'code': code, 'name': name, 'parent_code': parent_code, 'sort_order': sort_order, 'row': i})

    created, updated, errors = 0, 0, []
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            # 기존 code → id 매핑
            cur.execute("SELECT id, code FROM organizations")
            code_to_id = {r['code']: r['id'] for r in cur.fetchall()}

            # 1차: upsert (parent는 나중에)
            for rd in rows_data:
                try:
                    if rd['code'] in code_to_id:
                        cur.execute(
                            "UPDATE organizations SET name=%s, sort_order=%s WHERE code=%s",
                            (rd['name'], rd['sort_order'], rd['code'])
                        )
                        updated += 1
                    else:
                        cur.execute(
                            "INSERT INTO organizations (code, name, sort_order) VALUES (%s,%s,%s)",
                            (rd['code'], rd['name'], rd['sort_order'])
                        )
                        code_to_id[rd['code']] = cur.lastrowid
                        created += 1
                except Exception as e:
                    errors.append({'row': rd['row'], 'error': str(e)})

            # 2차: parent 설정
            for rd in rows_data:
                if rd['parent_code'] and rd['parent_code'] in code_to_id:
                    parent_id = code_to_id[rd['parent_code']]
                    cur.execute(
                        "UPDATE organizations SET parent_id=%s WHERE code=%s",
                        (parent_id, rd['code'])
                    )

    return HandlerResult(status=200, body={
        'created': created, 'updated': updated, 'errors': errors
    })


CIMS_ORG_HANDLER_LIST = [
    (_ORG_BASE, handle_organizations, {}),
]
