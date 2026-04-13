"""
CIMS Admin REST API
Subscriber and PTT group CRUD operations backed by MariaDB.

Routes (prefix-matched):
  /api/v1/users                               GET list / POST create
  /api/v1/users/{pid}                         GET / PUT / DELETE
  /api/v1/users/{pid}/call                    GET list / POST add call subscription
  /api/v1/users/{pid}/call/{msisdn}           PUT update / DELETE remove call subscription
  /api/v1/users/{pid}/ptt                     GET list / POST add PTT subscription
  /api/v1/users/{pid}/ptt/{msisdn}            PUT update / DELETE remove PTT subscription
  /api/v1/ptt/groups                          GET list / POST create
  /api/v1/ptt/groups/{id}                     GET / PUT / DELETE
  /api/v1/ptt/groups/{id}/members             GET list / POST add
  /api/v1/ptt/groups/{id}/members/{uid}       DELETE
"""

from urllib.parse import urlparse, unquote
from pathlib import PurePath

import pymysql
import pymysql.cursors

from util.pi_http.http_handler import HandlerArgs, HandlerResult
from csc_service import notify_csp

# ──────────────────────────────────────────────────────────────
#  DB helper
# ──────────────────────────────────────────────────────────────

def _get_db(config: dict):
    db = config.get('CimsDatabase', {})
    return pymysql.connect(
        host=db.get('Host', '127.0.0.1'),
        port=int(db.get('Port', 3306)),
        user=db.get('User', 'root'),
        password=db.get('Password', ''),
        database=db.get('Db', 'cims'),
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
    )


def _path_parts(full_path: str, base: str):
    """Return the path segments that come after *base*, URL-decoded.

    e.g. base='/api/v1/users', full_path='/api/v1/users/%2B821001' → ('+821001',)
    """
    path = urlparse(full_path).path
    try:
        rel = PurePath(path).relative_to(PurePath(base))
        return tuple(unquote(p) for p in rel.parts)
    except ValueError:
        return ()


def _dt(val):
    """Convert datetime to ISO string or None."""
    if val is None:
        return None
    return val.isoformat()


# ──────────────────────────────────────────────────────────────
#  Users handler
# ──────────────────────────────────────────────────────────────

_USERS_BASE = '/api/v1/users'


async def handle_users(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    config = kwargs.get('config', {})
    parts = _path_parts(handler_args.full_path, _USERS_BASE)
    # parts: () | (pid,) | (pid, 'call'|'ptt') | (pid, 'call'|'ptt', msisdn)
    person_id = parts[0] if len(parts) > 0 else None
    sub       = parts[1] if len(parts) > 1 else None   # 'call' | 'ptt'
    sub_id    = parts[2] if len(parts) > 2 else None   # MSISDN of the subscription
    method    = handler_args.method.upper()

    try:
        if person_id is None:
            if method == 'GET':
                return await _list_users(config)
            elif method == 'POST':
                return await _create_user(handler_args.body, config)
            return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        if person_id == 'batch' and method == 'DELETE':
            return await _batch_delete_users(handler_args.body, config)

        if sub is None:
            if method == 'GET':
                return await _get_user(person_id, config)
            elif method == 'PUT':
                return await _update_user(person_id, handler_args.body, config)
            elif method == 'DELETE':
                return await _delete_user(person_id, config)
            return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        if sub in ('call', 'ptt'):
            if sub_id is None:
                if method == 'GET':
                    return await _list_subscriptions(person_id, sub, config)
                elif method == 'POST':
                    return await _add_subscription(person_id, sub, handler_args.body, config)
                return HandlerResult(status=405, body={'error': 'Method Not Allowed'})
            else:
                if method == 'PUT':
                    return await _update_subscription(person_id, sub, sub_id, handler_args.body, config)
                elif method == 'DELETE':
                    return await _delete_subscription(person_id, sub, sub_id, config)
                return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        return HandlerResult(status=404, body={'error': 'Not Found'})
    except pymysql.Error as e:
        return HandlerResult(status=500, body={'error': str(e)})


def _has_email_column(cur) -> bool:
    """Check whether users.email column exists (migration may not have run yet)."""
    cur.execute(
        "SELECT COUNT(*) AS cnt FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='users' AND COLUMN_NAME='email'"
    )
    return cur.fetchone()['cnt'] > 0


async def _list_users(config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            has_email = _has_email_column(cur)
            email_col = ", u.email" if has_email else ""
            cur.execute(
                f"SELECT u.id, u.name, u.login_id{email_col}, u.org_id, u.details, "
                "u.create_time, u.update_time "
                "FROM users u "
                "ORDER BY u.id"
            )
            rows = cur.fetchall()
            for row in rows:
                if not has_email:
                    row['email'] = ''
                row['create_time'] = _dt(row['create_time'])
                row['update_time'] = _dt(row['update_time'])
                # attach reject list
                cur.execute(
                    "SELECT reject_id FROM user_rejects WHERE user_id=%s",
                    (row['id'],)
                )
                row['reject_id'] = [r['reject_id'] for r in cur.fetchall()]
                # attach subscriptions
                cur.execute(
                    "SELECT id, auth_id, dnd, forward_id, register_time, logout_time "
                    "FROM voip_subscriptions WHERE user_id=%s ORDER BY id",
                    (row['id'],)
                )
                call_subs = cur.fetchall()
                for s in call_subs:
                    s['dnd'] = bool(s['dnd'])
                    s['register_time'] = _dt(s['register_time'])
                    s['logout_time']   = _dt(s['logout_time'])
                row['call_subscriptions'] = call_subs

                cur.execute(
                    "SELECT id, auth_id, dnd, forward_id, register_time, logout_time "
                    "FROM ptt_subscriptions WHERE user_id=%s ORDER BY id",
                    (row['id'],)
                )
                ptt_subs = cur.fetchall()
                for s in ptt_subs:
                    s['dnd'] = bool(s['dnd'])
                    s['register_time'] = _dt(s['register_time'])
                    s['logout_time']   = _dt(s['logout_time'])
                row['ptt_subscriptions'] = ptt_subs
    return HandlerResult(status=200, body={'users': rows})


async def _get_user(person_id: str, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            has_email = _has_email_column(cur)
            email_col = ", email" if has_email else ""
            cur.execute(
                f"SELECT id, name{email_col}, org_id, details, create_time, update_time "
                "FROM users WHERE id=%s",
                (person_id,)
            )
            row = cur.fetchone()
            if row is None:
                return HandlerResult(status=404, body={'error': 'User not found'})
            if not has_email:
                row['email'] = ''
            row['create_time'] = _dt(row['create_time'])
            row['update_time'] = _dt(row['update_time'])

            # reject list
            cur.execute(
                "SELECT reject_id FROM user_rejects WHERE user_id=%s",
                (person_id,)
            )
            row['reject_id'] = [r['reject_id'] for r in cur.fetchall()]

            # call subscriptions
            cur.execute(
                "SELECT id, auth_id, dnd, forward_id, register_time, logout_time "
                "FROM voip_subscriptions WHERE user_id=%s ORDER BY id",
                (person_id,)
            )
            call_subs = cur.fetchall()
            for s in call_subs:
                s['dnd'] = bool(s['dnd'])
                s['register_time'] = _dt(s['register_time'])
                s['logout_time']   = _dt(s['logout_time'])
            row['call_subscriptions'] = call_subs

            # ptt subscriptions
            cur.execute(
                "SELECT id, auth_id, dnd, forward_id, register_time, logout_time "
                "FROM ptt_subscriptions WHERE user_id=%s ORDER BY id",
                (person_id,)
            )
            ptt_subs = cur.fetchall()
            for s in ptt_subs:
                s['dnd'] = bool(s['dnd'])
                s['register_time'] = _dt(s['register_time'])
                s['logout_time']   = _dt(s['logout_time'])
            row['ptt_subscriptions'] = ptt_subs

    return HandlerResult(status=200, body=row)


async def _create_user(body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})
    name = body.get('name', '').strip()
    if not name:
        return HandlerResult(status=400, body={'error': 'name is required'})

    login_id   = body.get('login_id', '').strip()
    password   = body.get('password', '')
    email      = body.get('email', '')
    org_id     = body.get('org_id', '')
    details    = body.get('details') or None
    reject_ids = body.get('reject_id', [])

    # login_id 미지정 시 name 기반 자동 생성
    if not login_id:
        login_id = name.replace(' ', '_').lower()

    # password → SHA-256 해시
    import hashlib
    pw_hash = hashlib.sha256(password.encode()).hexdigest() if password else ''

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            has_email = _has_email_column(cur)
            if has_email:
                cur.execute(
                    "INSERT INTO users "
                    "(name, login_id, password, email, org_id, details, create_time, update_time) "
                    "VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())",
                    (name, login_id, pw_hash, email, org_id, details)
                )
            else:
                cur.execute(
                    "INSERT INTO users "
                    "(name, login_id, password, org_id, details, create_time, update_time) "
                    "VALUES (%s, %s, %s, %s, %s, NOW(), NOW())",
                    (name, login_id, pw_hash, org_id, details)
                )
            person_id = cur.lastrowid

            if reject_ids:
                for rid in reject_ids:
                    cur.execute(
                        "INSERT IGNORE INTO user_rejects (user_id, reject_id) VALUES (%s, %s)",
                        (person_id, rid)
                    )
    return HandlerResult(status=201, body={'id': person_id})


async def _update_user(person_id: str, body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})

    fields = []
    values = []
    for col in ('name', 'email', 'org_id', 'details'):
        if col in body:
            fields.append(f'{col}=%s')
            values.append(body[col])

    if fields:
        fields.append('update_time=NOW()')
        values.append(person_id)
        with _get_db(config) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE users SET {', '.join(fields)} WHERE id=%s",
                    values
                )
                if cur.rowcount == 0:
                    return HandlerResult(status=404, body={'error': 'User not found'})
                if 'reject_id' in body:
                    cur.execute("DELETE FROM user_rejects WHERE user_id=%s", (person_id,))
                    for rid in body['reject_id']:
                        cur.execute(
                            "INSERT IGNORE INTO user_rejects (user_id, reject_id) VALUES (%s, %s)",
                            (person_id, rid)
                        )
    elif 'reject_id' in body:
        with _get_db(config) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM users WHERE id=%s", (person_id,))
                if cur.fetchone() is None:
                    return HandlerResult(status=404, body={'error': 'User not found'})
                cur.execute("DELETE FROM user_rejects WHERE user_id=%s", (person_id,))
                for rid in body['reject_id']:
                    cur.execute(
                        "INSERT IGNORE INTO user_rejects (user_id, reject_id) VALUES (%s, %s)",
                        (person_id, rid)
                    )
    else:
        return HandlerResult(status=400, body={'error': 'No updatable fields provided'})

    return HandlerResult(status=200, body={'id': person_id})


async def _delete_user(person_id: str, config):
    sub_ids = []
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            # 삭제 전 연관 subscription ID 수집
            for table in ('voip_subscriptions', 'ptt_subscriptions'):
                cur.execute(f"SELECT id FROM {table} WHERE user_id=%s", (person_id,))
                sub_ids.extend(r['id'] for r in cur.fetchall())
            cur.execute("DELETE FROM voip_subscriptions WHERE user_id=%s", (person_id,))
            cur.execute("DELETE FROM ptt_subscriptions WHERE user_id=%s", (person_id,))
            cur.execute("DELETE FROM user_rejects WHERE user_id=%s", (person_id,))
            cur.execute("DELETE FROM users WHERE id=%s", (person_id,))
            if cur.rowcount == 0:
                return HandlerResult(status=404, body={'error': 'User not found'})
    for sid in sub_ids:
        notify_csp("USER_CHANGED", f"tel:{sid}", "DELETE")
    return HandlerResult(status=200, body={'id': person_id})


async def _batch_delete_users(body, config):
    """다수의 가입자를 한번에 삭제"""
    ids = body.get('ids', []) if body else []
    if not ids:
        return HandlerResult(status=400, body={'error': 'ids 필드가 필요합니다'})

    deleted = 0
    errors = []
    for pid in ids:
        try:
            result = await _delete_user(str(pid), config)
            if result.status == 200:
                deleted += 1
            else:
                errors.append({'id': pid, 'error': 'not found'})
        except Exception as e:
            errors.append({'id': pid, 'error': str(e)})

    return HandlerResult(status=200, body={
        'deleted': deleted, 'errors': errors
    })


# ──────────────────────────────────────────────────────────────
#  Subscription handlers (call / ptt)
# ──────────────────────────────────────────────────────────────

def _sub_table(svc: str) -> str:
    return 'voip_subscriptions' if svc == 'call' else 'ptt_subscriptions'




async def _list_subscriptions(person_id: str, svc: str, config):
    table = _sub_table(svc)
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM users WHERE id=%s", (person_id,))
            if cur.fetchone() is None:
                return HandlerResult(status=404, body={'error': 'User not found'})
            cur.execute(
                f"SELECT id, auth_id, dnd, forward_id, register_time, logout_time "
                f"FROM {table} WHERE user_id=%s ORDER BY id",
                (person_id,)
            )
            subs = cur.fetchall()
            for s in subs:
                s['dnd'] = bool(s['dnd'])
                s['register_time'] = _dt(s['register_time'])
                s['logout_time']   = _dt(s['logout_time'])
    return HandlerResult(status=200, body={'subscriptions': subs})


async def _add_subscription(person_id: str, svc: str, body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})
    msisdn = body.get('id', '').strip()
    if not msisdn:
        return HandlerResult(status=400, body={'error': 'id (MSISDN) is required'})

    auth_id    = body.get('auth_id', msisdn)
    passwd     = body.get('passwd', '')
    dnd        = 1 if body.get('dnd', False) else 0
    forward_id = body.get('forward_id', '')
    table      = _sub_table(svc)

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM users WHERE id=%s", (person_id,))
            if cur.fetchone() is None:
                return HandlerResult(status=404, body={'error': 'User not found'})
            cur.execute(
                f"INSERT INTO {table} (id, user_id, auth_id, passwd, dnd, forward_id) "
                f"VALUES (%s, %s, %s, %s, %s, %s)",
                (msisdn, person_id, auth_id, passwd, dnd, forward_id)
            )
    notify_csp("USER_CHANGED", f"tel:{msisdn}", "POST")
    return HandlerResult(status=201, body={'id': msisdn})


async def _update_subscription(person_id: str, svc: str, msisdn: str, body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})

    auth_id    = body.get('auth_id', msisdn)
    passwd     = body.get('passwd', '')
    dnd        = 1 if body.get('dnd', False) else 0
    forward_id = body.get('forward_id', '')
    table      = _sub_table(svc)

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE {table} SET auth_id=%s, passwd=%s, dnd=%s, forward_id=%s "
                f"WHERE id=%s AND user_id=%s",
                (auth_id, passwd, dnd, forward_id, msisdn, person_id)
            )
            if cur.rowcount == 0:
                return HandlerResult(status=404, body={'error': 'Subscription not found'})
    notify_csp("USER_CHANGED", f"tel:{msisdn}", "PUT")
    return HandlerResult(status=200, body={'id': msisdn})


async def _delete_subscription(person_id: str, svc: str, msisdn: str, config):
    table = _sub_table(svc)
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"DELETE FROM {table} WHERE id=%s AND user_id=%s",
                (msisdn, person_id)
            )
            if cur.rowcount == 0:
                return HandlerResult(status=404, body={'error': 'Subscription not found'})
    notify_csp("USER_CHANGED", f"tel:{msisdn}", "DELETE")
    return HandlerResult(status=200, body={'id': msisdn})


# ──────────────────────────────────────────────────────────────
#  PTT Groups handler
# ──────────────────────────────────────────────────────────────

_GROUPS_BASE = '/api/v1/ptt/groups'


async def handle_ptt_groups(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    """Handles all /api/v1/ptt/groups/* routes.

    Path structure:
      /api/v1/ptt/groups                   → group_id=None
      /api/v1/ptt/groups/{id}              → group_id set, sub=None
      /api/v1/ptt/groups/{id}/members      → sub='members', member_id=None
      /api/v1/ptt/groups/{id}/members/{uid}→ sub='members', member_id set
    """
    config = kwargs.get('config', {})
    parts = _path_parts(handler_args.full_path, _GROUPS_BASE)
    group_id  = parts[0] if len(parts) > 0 else None
    sub       = parts[1] if len(parts) > 1 else None   # 'members'
    member_id = parts[2] if len(parts) > 2 else None
    method = handler_args.method.upper()

    try:
        if group_id is None:
            if method == 'GET':
                return await _list_groups(config)
            elif method == 'POST':
                return await _create_group(handler_args.body, config)
            return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        if sub is None:
            if method == 'GET':
                return await _get_group(group_id, config)
            elif method == 'PUT':
                return await _update_group(group_id, handler_args.body, config)
            elif method == 'DELETE':
                return await _delete_group(group_id, config)
            return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        if sub == 'members':
            if member_id is None:
                if method == 'GET':
                    return await _list_members(group_id, config)
                elif method == 'POST':
                    return await _add_member(group_id, handler_args.body, config)
                return HandlerResult(status=405, body={'error': 'Method Not Allowed'})
            else:
                if method == 'DELETE':
                    return await _remove_member(group_id, member_id, config)
                return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

        return HandlerResult(status=404, body={'error': 'Not Found'})
    except pymysql.Error as e:
        return HandlerResult(status=500, body={'error': str(e)})


async def _list_groups(config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, name, video_enabled, priority, encryption, emergency_call, "
                "org_code, session_start, session_end FROM ptt_groups ORDER BY id"
            )
            groups = cur.fetchall()
            for g in groups:
                g['video_enabled'] = bool(g.get('video_enabled', 0))
                g['encryption'] = bool(g.get('encryption', 0))
                g['emergency_call'] = bool(g.get('emergency_call', 0))
                if g.get('session_start'): g['session_start'] = g['session_start'].isoformat()
                if g.get('session_end'): g['session_end'] = g['session_end'].isoformat()
                cur.execute(
                    "SELECT user_id, priority FROM ptt_group_members "
                    "WHERE group_id=%s ORDER BY priority",
                    (g['id'],)
                )
                g['members'] = cur.fetchall()
    return HandlerResult(status=200, body={'groups': groups})


async def _get_group(group_id: str, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, name, video_enabled, priority, encryption, emergency_call, "
                "org_code, session_start, session_end FROM ptt_groups WHERE id=%s",
                (group_id,)
            )
            group = cur.fetchone()
            if group is None:
                return HandlerResult(status=404, body={'error': 'Group not found'})
            group['video_enabled'] = bool(group.get('video_enabled', 0))
            group['encryption'] = bool(group.get('encryption', 0))
            group['emergency_call'] = bool(group.get('emergency_call', 0))
            if group.get('session_start'): group['session_start'] = group['session_start'].isoformat()
            if group.get('session_end'): group['session_end'] = group['session_end'].isoformat()
            cur.execute(
                "SELECT user_id, priority FROM ptt_group_members "
                "WHERE group_id=%s ORDER BY priority",
                (group_id,)
            )
            group['members'] = cur.fetchall()
    return HandlerResult(status=200, body=group)


async def _create_group(body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})
    group_id = body.get('id', '').strip()
    if not group_id:
        return HandlerResult(status=400, body={'error': 'id is required'})
    name           = body.get('name', group_id)
    video_enabled  = 1 if body.get('video_enabled', False) else 0
    priority       = int(body.get('priority', 5))
    encryption     = 1 if body.get('encryption', False) else 0
    emergency_call = 1 if body.get('emergency_call', False) else 0
    org_code       = body.get('org_code', '') or None
    session_start  = body.get('session_start') or None
    session_end    = body.get('session_end') or None
    members        = body.get('members', [])

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO ptt_groups (id, name, video_enabled, priority, encryption, "
                "emergency_call, org_code, session_start, session_end) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (group_id, name, video_enabled, priority, encryption,
                 emergency_call, org_code, session_start, session_end)
            )
            for m in members:
                uid  = m.get('user_id', m.get('id', ''))
                prio = int(m.get('priority', 0))
                if uid:
                    cur.execute(
                        "INSERT IGNORE INTO ptt_group_members "
                        "(group_id, user_id, priority) VALUES (%s, %s, %s)",
                        (group_id, uid, prio)
                    )
    notify_csp("GROUP_CHANGED", f"tel:{group_id}", "POST")
    return HandlerResult(status=201, body={'id': group_id})


async def _update_group(group_id: str, body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            update_fields = []
            update_vals   = []
            if 'name' in body:
                update_fields.append('name=%s')
                update_vals.append(body['name'])
            if 'video_enabled' in body:
                update_fields.append('video_enabled=%s')
                update_vals.append(1 if body['video_enabled'] else 0)
            for fld in ('priority',):
                if fld in body:
                    update_fields.append(f'{fld}=%s')
                    update_vals.append(int(body[fld]))
            for fld in ('encryption', 'emergency_call'):
                if fld in body:
                    update_fields.append(f'{fld}=%s')
                    update_vals.append(1 if body[fld] else 0)
            for fld in ('org_code',):
                if fld in body:
                    update_fields.append(f'{fld}=%s')
                    update_vals.append(body[fld] or None)
            for fld in ('session_start', 'session_end'):
                if fld in body:
                    update_fields.append(f'{fld}=%s')
                    update_vals.append(body[fld] or None)
            if update_fields:
                update_vals.append(group_id)
                cur.execute(
                    "UPDATE ptt_groups SET " + ", ".join(update_fields) + " WHERE id=%s",
                    update_vals
                )
                if cur.rowcount == 0:
                    return HandlerResult(status=404, body={'error': 'Group not found'})
            if 'members' in body:
                cur.execute(
                    "DELETE FROM ptt_group_members WHERE group_id=%s",
                    (group_id,)
                )
                for m in body['members']:
                    uid  = m.get('user_id', m.get('id', ''))
                    prio = int(m.get('priority', 0))
                    if uid:
                        cur.execute(
                            "INSERT IGNORE INTO ptt_group_members "
                            "(group_id, user_id, priority) VALUES (%s, %s, %s)",
                            (group_id, uid, prio)
                        )
    notify_csp("GROUP_CHANGED", f"tel:{group_id}", "PUT")
    return HandlerResult(status=200, body={'id': group_id})


async def _delete_group(group_id: str, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM ptt_group_members WHERE group_id=%s",
                (group_id,)
            )
            cur.execute(
                "DELETE FROM ptt_groups WHERE id=%s",
                (group_id,)
            )
            if cur.rowcount == 0:
                return HandlerResult(status=404, body={'error': 'Group not found'})
    notify_csp("GROUP_CHANGED", f"tel:{group_id}", "DELETE")
    return HandlerResult(status=200, body={'id': group_id})


async def _list_members(group_id: str, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM ptt_groups WHERE id=%s", (group_id,))
            if cur.fetchone() is None:
                return HandlerResult(status=404, body={'error': 'Group not found'})
            cur.execute(
                "SELECT user_id, priority FROM ptt_group_members "
                "WHERE group_id=%s ORDER BY priority",
                (group_id,)
            )
            members = cur.fetchall()
    return HandlerResult(status=200, body={'group_id': group_id, 'members': members})


async def _add_member(group_id: str, body, config):
    if not isinstance(body, dict):
        return HandlerResult(status=400, body={'error': 'JSON body required'})
    user_id  = body.get('user_id', '').strip()
    if not user_id:
        return HandlerResult(status=400, body={'error': 'user_id is required'})
    priority = int(body.get('priority', 0))

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM ptt_groups WHERE id=%s", (group_id,))
            if cur.fetchone() is None:
                return HandlerResult(status=404, body={'error': 'Group not found'})
            cur.execute(
                "INSERT INTO ptt_group_members (group_id, user_id, priority) "
                "VALUES (%s, %s, %s) "
                "ON DUPLICATE KEY UPDATE priority=VALUES(priority)",
                (group_id, user_id, priority)
            )
    notify_csp("GROUP_CHANGED", f"tel:{group_id}", "PUT")
    return HandlerResult(status=201, body={'group_id': group_id, 'user_id': user_id})


async def _remove_member(group_id: str, user_id: str, config):
    with _get_db(config) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM ptt_group_members WHERE group_id=%s AND user_id=%s",
                (group_id, user_id)
            )
            if cur.rowcount == 0:
                return HandlerResult(status=404, body={'error': 'Member not found'})
    notify_csp("GROUP_CHANGED", f"tel:{group_id}", "PUT")
    return HandlerResult(status=200, body={'group_id': group_id, 'user_id': user_id})


# ──────────────────────────────────────────────────────────────
#  Handler list (registered in app.py)
# ──────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────
#  Call logs handler
# ──────────────────────────────────────────────────────────────

_CALL_LOGS_BASE = '/api/v1/call/logs'

_END_REASON_KO = {
    'normal': '정상종료', 'busy': '통화중', 'cancel': '취소',
    'timeout': '시간초과', 'error': '오류',
}


def _call_log_row(row: dict) -> dict:
    row['invite_time'] = _dt(row['invite_time'])
    row['answer_time'] = _dt(row['answer_time'])
    row['end_time']    = _dt(row['end_time'])
    row['end_reason_ko'] = _END_REASON_KO.get(row.get('end_reason') or '', '')
    return row


def _tables_exist(cur) -> bool:
    cur.execute(
        "SELECT COUNT(*) AS cnt FROM information_schema.TABLES "
        "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='voip_call_logs'"
    )
    return cur.fetchone()['cnt'] > 0


async def handle_call_logs(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    """
    GET /api/v1/call/logs
        ?state=ringing|active|ended   (default: all)
        ?caller=MSISDN
        ?callee=MSISDN
        ?msisdn=MSISDN                (caller OR callee OR participant)
        ?group_id=ID                  (PTT only)
        ?call_type=voip|ptt
        ?from_dt=YYYY-MM-DD
        ?to_dt=YYYY-MM-DD
        ?limit=N                      (default 200, max 1000)
        ?offset=N                     (default 0)

    GET /api/v1/call/logs/active      (shorthand: state=ringing,active)
    """
    from urllib.parse import urlparse, parse_qs, unquote

    config = kwargs.get('config', {})
    parsed = urlparse(handler_args.full_path)
    qs     = parse_qs(parsed.query)

    def qp(name, default=None):
        vals = qs.get(name)
        return unquote(vals[0]) if vals else default

    # distinguish /call/logs vs /call/logs/active
    path_tail = parsed.path.rstrip('/')
    active_only = path_tail.endswith('/active')

    if handler_args.method.upper() != 'GET':
        return HandlerResult(status=405, body={'error': 'Method Not Allowed'})

    state     = qp('state')
    caller    = qp('caller')
    callee    = qp('callee')
    msisdn    = qp('msisdn')
    group_id  = qp('group_id')
    call_type = qp('call_type')
    from_dt   = qp('from_dt')
    to_dt     = qp('to_dt')
    limit     = min(int(qp('limit', 200)), 1000)
    offset    = int(qp('offset', 0))

    if active_only:
        state = 'active_ringing'  # special token

    try:
        with _get_db(config) as conn:
            with conn.cursor() as cur:
                if not _tables_exist(cur):
                    return HandlerResult(status=503, body={
                        'error': 'Call log tables not created yet. Run: sudo mysql cims < sql/migrate_call_logs.sql'
                    })

                # ── build per-table WHERE clauses ──
                # voip_call_logs columns: id, call_id, initiator, callee, state, invite_time, ...
                # ptt_call_logs  columns: id, call_id, group_id, initiator, state, invite_time, ...

                def _build_where(for_ptt: bool):
                    w = []
                    p = []
                    if state == 'active_ringing':
                        w.append("l.state IN ('ringing','active')")
                    elif state:
                        w.append("l.state = %s")
                        p.append(state)
                    if caller:
                        w.append("l.initiator = %s")
                        p.append(caller)
                    if from_dt:
                        w.append("l.invite_time >= %s")
                        p.append(from_dt + ' 00:00:00')
                    if to_dt:
                        w.append("l.invite_time <= %s")
                        p.append(to_dt + ' 23:59:59')
                    if for_ptt:
                        if group_id:
                            w.append("l.group_id = %s")
                            p.append(group_id)
                        if msisdn:
                            w.append(
                                "(l.initiator = %s OR EXISTS("
                                " SELECT 1 FROM ptt_call_participants cp"
                                " WHERE cp.log_id = l.id AND cp.msisdn = %s))"
                            )
                            p += [msisdn, msisdn]
                    else:
                        if callee:
                            w.append("l.callee = %s")
                            p.append(callee)
                        if msisdn:
                            w.append(
                                "(l.initiator = %s OR l.callee = %s OR EXISTS("
                                " SELECT 1 FROM voip_call_participants cp"
                                " WHERE cp.log_id = l.id AND cp.msisdn = %s))"
                            )
                            p += [msisdn, msisdn, msisdn]
                    return ("WHERE " + " AND ".join(w)) if w else "", p

                # determine which tables to query
                use_voip = call_type != 'ptt'
                use_ptt  = call_type != 'voip'

                # ── build UNION query ──
                union_parts = []
                union_params = []

                if use_voip:
                    voip_where, voip_params = _build_where(for_ptt=False)
                    union_parts.append(
                        f"SELECT l.id, l.call_id, 'voip' AS call_type, "
                        f"NULL AS group_id, "
                        f"l.initiator, l.callee, l.state, "
                        f"l.invite_time, l.answer_time, l.end_time, "
                        f"l.duration, l.sip_status, l.end_reason "
                        f"FROM voip_call_logs l {voip_where}"
                    )
                    union_params += voip_params

                if use_ptt:
                    ptt_where, ptt_params = _build_where(for_ptt=True)
                    union_parts.append(
                        f"SELECT l.id, l.call_id, 'ptt' AS call_type, "
                        f"l.group_id, "
                        f"l.initiator, NULL AS callee, l.state, "
                        f"l.invite_time, l.answer_time, l.end_time, "
                        f"l.duration, NULL AS sip_status, l.end_reason "
                        f"FROM ptt_call_logs l {ptt_where}"
                    )
                    union_params += ptt_params

                union_sql = " UNION ALL ".join(union_parts)

                # ── total count ──
                cur.execute(
                    f"SELECT COUNT(*) AS cnt FROM ({union_sql}) AS combined",
                    union_params
                )
                total = cur.fetchone()['cnt']

                # ── main query ──
                cur.execute(
                    f"SELECT * FROM ({union_sql}) AS combined "
                    f"ORDER BY invite_time DESC "
                    f"LIMIT %s OFFSET %s",
                    union_params + [limit, offset]
                )
                logs = cur.fetchall()

                for row in logs:
                    _call_log_row(row)

                    # attach participants from the correct split table
                    parts_table = 'ptt_call_participants' if row['call_type'] == 'ptt' else 'voip_call_participants'
                    cur.execute(
                        f"SELECT msisdn, role, join_time, leave_time "
                        f"FROM {parts_table} WHERE log_id = %s "
                        f"ORDER BY role, join_time",
                        (row['id'],)
                    )
                    parts = cur.fetchall()
                    for p in parts:
                        p['join_time']  = _dt(p['join_time'])
                        p['leave_time'] = _dt(p['leave_time'])
                    row['participants'] = parts

        return HandlerResult(status=200, body={
            'total': total, 'limit': limit, 'offset': offset, 'logs': logs
        })
    except pymysql.Error as e:
        return HandlerResult(status=500, body={'error': str(e)})


# ──────────────────────────────────────────────────────────────
#  Excel Import / Template
# ──────────────────────────────────────────────────────────────

_IMPORT_BASE = '/api/v1/users/import'


async def handle_import(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    config = kwargs.get('config', {})
    method = handler_args.method.upper()
    parts = _path_parts(handler_args.full_path, _IMPORT_BASE)

    if len(parts) >= 1 and parts[0] == 'template' and method == 'GET':
        return _generate_template()

    if method == 'POST':
        return await _process_import(handler_args, config)

    return HandlerResult(status=405, body={'error': 'Method Not Allowed'})


def _generate_template():
    """빈 Excel 템플릿 생성 후 반환"""
    import openpyxl
    import io

    wb = openpyxl.Workbook()
    # Sheet 1: users
    ws1 = wb.active
    ws1.title = 'users'
    ws1.append(['name', 'login_id', 'org_code', 'details', 'reject_ids'])
    ws1.append(['홍길동', 'hong', 'DEV_01', '개발1팀', '+8210001,+8210002'])

    # Sheet 2: voip_subscriptions
    ws2 = wb.create_sheet('voip_subscriptions')
    ws2.append(['name', 'msisdn', 'auth_id', 'password', 'dnd', 'forward_id'])
    ws2.append(['홍길동', '+821357007100', '45003310000100@ims.domain', '123456', 'N', ''])

    # Sheet 3: ptt_subscriptions
    ws3 = wb.create_sheet('ptt_subscriptions')
    ws3.append(['name', 'msisdn', 'auth_id', 'password', 'dnd'])
    ws3.append(['홍길동', '+82571900100', '', '123456', 'N'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return HandlerResult(status=200, body=buf.getvalue(), headers={
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename="cims_import_template.xlsx"',
    })


async def _process_import(handler_args: HandlerArgs, config):
    """Excel 파일을 파싱하여 가입자/구독 일괄 등록"""
    import openpyxl
    import io

    # multipart body에서 파일 데이터 추출
    file_data = handler_args.raw_body if hasattr(handler_args, 'raw_body') else None
    if not file_data:
        # JSON body에 base64 인코딩된 파일이 올 수도 있음
        body = handler_args.body or {}
        if 'file_base64' in body:
            import base64
            file_data = base64.b64decode(body['file_base64'])
        else:
            return HandlerResult(status=400, body={'error': 'file_base64 필드가 필요합니다'})

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data))
    except Exception as e:
        return HandlerResult(status=400, body={'error': f'Excel 파일 파싱 실패: {e}'})

    result = {'created_users': 0, 'created_voip': 0, 'created_ptt': 0, 'errors': []}
    name_to_id = {}  # name → person_id 매핑

    with _get_db(config) as conn:
        with conn.cursor() as cur:
            # 기존 사용자 name → id 매핑 로드
            cur.execute("SELECT id, name FROM users")
            for row in cur.fetchall():
                name_to_id[row['name']] = row['id']

            # Sheet 1: users
            if 'users' in wb.sheetnames:
                ws = wb['users']
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    rd = dict(zip(headers, row))
                    name = str(rd.get('name', '') or '').strip()
                    if not name:
                        result['errors'].append({'row': i, 'sheet': 'users', 'error': 'name 필수'})
                        continue
                    if name in name_to_id:
                        continue  # 이미 존재, 스킵

                    login_id = str(rd.get('login_id', '') or '').strip()
                    org_id = str(rd.get('org_code', '') or '').strip()
                    details = str(rd.get('details', '') or '').strip()
                    reject_ids = str(rd.get('reject_ids', '') or '').strip()

                    try:
                        cur.execute(
                            "INSERT INTO users (name, login_id, org_id, details) VALUES (%s,%s,%s,%s)",
                            (name, login_id, org_id, details)
                        )
                        pid = cur.lastrowid
                        name_to_id[name] = pid
                        result['created_users'] += 1

                        if reject_ids:
                            for rid in reject_ids.split(','):
                                rid = rid.strip()
                                if rid:
                                    cur.execute(
                                        "INSERT IGNORE INTO user_rejects (user_id, reject_id) VALUES (%s,%s)",
                                        (pid, rid)
                                    )
                    except Exception as e:
                        result['errors'].append({'row': i, 'sheet': 'users', 'error': str(e)})

            # Sheet 2: voip_subscriptions
            if 'voip_subscriptions' in wb.sheetnames:
                ws = wb['voip_subscriptions']
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    rd = dict(zip(headers, row))
                    name = str(rd.get('name', '') or '').strip()
                    msisdn = str(rd.get('msisdn', '') or '').strip()
                    if not name or not msisdn:
                        result['errors'].append({'row': i, 'sheet': 'voip_subscriptions', 'error': 'name/msisdn 필수'})
                        continue

                    pid = name_to_id.get(name)
                    if not pid:
                        # 자동 생성
                        try:
                            cur.execute("INSERT INTO users (name) VALUES (%s)", (name,))
                            pid = cur.lastrowid
                            name_to_id[name] = pid
                            result['created_users'] += 1
                        except Exception as e:
                            result['errors'].append({'row': i, 'sheet': 'voip_subscriptions', 'error': f'사용자 생성 실패: {e}'})
                            continue

                    auth_id = str(rd.get('auth_id', '') or '').strip() or msisdn
                    passwd = str(rd.get('password', '') or '').strip() or '123456'
                    dnd = 1 if str(rd.get('dnd', '')).upper() in ('Y', 'YES', '1', 'TRUE') else 0
                    forward_id = str(rd.get('forward_id', '') or '').strip()

                    try:
                        cur.execute(
                            "INSERT IGNORE INTO voip_subscriptions (id, user_id, auth_id, passwd, dnd, forward_id) "
                            "VALUES (%s,%s,%s,%s,%s,%s)",
                            (msisdn, pid, auth_id, passwd, dnd, forward_id)
                        )
                        if cur.rowcount > 0:
                            result['created_voip'] += 1
                            notify_csp("USER_CHANGED", f"tel:{msisdn}", "PUT")
                        else:
                            result['errors'].append({'row': i, 'sheet': 'voip_subscriptions', 'error': f'MSISDN 중복: {msisdn}'})
                    except Exception as e:
                        result['errors'].append({'row': i, 'sheet': 'voip_subscriptions', 'error': str(e)})

            # Sheet 3: ptt_subscriptions
            if 'ptt_subscriptions' in wb.sheetnames:
                ws = wb['ptt_subscriptions']
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    rd = dict(zip(headers, row))
                    name = str(rd.get('name', '') or '').strip()
                    msisdn = str(rd.get('msisdn', '') or '').strip()
                    if not name or not msisdn:
                        result['errors'].append({'row': i, 'sheet': 'ptt_subscriptions', 'error': 'name/msisdn 필수'})
                        continue

                    pid = name_to_id.get(name)
                    if not pid:
                        try:
                            cur.execute("INSERT INTO users (name) VALUES (%s)", (name,))
                            pid = cur.lastrowid
                            name_to_id[name] = pid
                            result['created_users'] += 1
                        except Exception as e:
                            result['errors'].append({'row': i, 'sheet': 'ptt_subscriptions', 'error': f'사용자 생성 실패: {e}'})
                            continue

                    auth_id = str(rd.get('auth_id', '') or '').strip() or msisdn
                    passwd = str(rd.get('password', '') or '').strip() or '123456'
                    dnd = 1 if str(rd.get('dnd', '')).upper() in ('Y', 'YES', '1', 'TRUE') else 0

                    try:
                        cur.execute(
                            "INSERT IGNORE INTO ptt_subscriptions (id, user_id, auth_id, passwd, dnd) "
                            "VALUES (%s,%s,%s,%s,%s)",
                            (msisdn, pid, auth_id, passwd, dnd)
                        )
                        if cur.rowcount > 0:
                            result['created_ptt'] += 1
                            notify_csp("USER_CHANGED", f"tel:{msisdn}", "PUT")
                        else:
                            result['errors'].append({'row': i, 'sheet': 'ptt_subscriptions', 'error': f'MSISDN 중복: {msisdn}'})
                    except Exception as e:
                        result['errors'].append({'row': i, 'sheet': 'ptt_subscriptions', 'error': str(e)})

    result['total'] = result['created_users'] + result['created_voip'] + result['created_ptt']
    return HandlerResult(status=200, body=result)


CIMS_ADMIN_HANDLER_LIST = [
    (_USERS_BASE,    handle_users,      {}),
    (_IMPORT_BASE,   handle_import,     {}),
    (_GROUPS_BASE,   handle_ptt_groups, {}),
    # call_logs는 csc_flow.py의 파일시스템 기반 API로 이동
]
