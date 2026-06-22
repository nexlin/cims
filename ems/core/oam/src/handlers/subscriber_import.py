"""가입자/조직 Excel 일괄 등록 오케스트레이션 (oam-svc).

Excel 파싱(openpyxl)은 운영 프로비저닝 툴링 — csc(headless 데이터/규격 API)가 아니라
**oam-svc(운영 오케스트레이션 평면)** 에 둔다. oam-svc 가 업로드된 Excel 을 파싱하고, 수신한
Bearer 토큰을 그대로 전달해 **csc JSON API**(POST /api/v1/users, /users/{id}/call|ptt,
/organizations …)를 호출한다. csc DB 직접 접근 없음 — 계약(HTTP) 경유.

라우트 (게이트웨이가 csc 의 /users·/organizations 보다 더 구체적인 세그먼트로 oam-svc 에 우선 매핑):
  GET  /api/v1/users/import/template          빈 가입자 Excel 템플릿 (users/volte/ptt 시트)
  POST /api/v1/users/import                   가입자/구독 일괄 등록 (file_base64)
  GET  /api/v1/organizations/import/template  빈 조직 Excel 템플릿
  POST /api/v1/organizations/import           조직 일괄 등록 (file_base64)

RBAC: manager+ (변경 작업). 템플릿 GET 도 manager+ (csc 정책과 동일).
"""

import asyncio
import base64
import io
from urllib.parse import urlparse, unquote
from pathlib import PurePath

import requests
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass

from httpsrv.handler import HandlerArgs, HandlerResult
from services import admin_auth, logger

_USERS_IMPORT_BASE = '/api/v1/users/import'
_ORG_IMPORT_BASE = '/api/v1/organizations/import'

_XLSX_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


# ── csc 업스트림 (loopback) ─────────────────────────────────────────────
def _csc_base(config: dict) -> str:
    # 게이트웨이·csc 동일 호스트 → loopback. config 로 override 가능(Subscriber.CscUrl).
    sub = config.get('Subscriber') if isinstance(config.get('Subscriber'), dict) else {}
    return (sub.get('CscUrl') or config.get('CscUrl') or 'https://127.0.0.1:4421').rstrip('/')


def _bearer(handler_args: HandlerArgs) -> str:
    for k, v in (handler_args.headers or {}).items():
        if k.lower() == 'authorization' and isinstance(v, str):
            return v.split(None, 1)[1] if v.lower().startswith('bearer ') else v
    return ''


def _hdr(token: str) -> dict:
    return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}


def _get(base, path, token):
    return requests.get(base + path, headers=_hdr(token), verify=False, timeout=15)


def _post(base, path, token, body):
    return requests.post(base + path, headers=_hdr(token), json=body, verify=False, timeout=15)


def _put(base, path, token, body):
    return requests.put(base + path, headers=_hdr(token), json=body, verify=False, timeout=15)


def _cell_str(v) -> str:
    return str(v).strip() if v not in (None, '') else ''


# ── 핸들러 진입 ─────────────────────────────────────────────────────────
async def handle_users_import(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    config = kwargs.get('config', {})
    _, err = admin_auth.require_role(handler_args, 'manager')
    if err:
        return err
    parts = _path_parts(handler_args.full_path, _USERS_IMPORT_BASE)
    method = handler_args.method.upper()
    if parts and parts[0] == 'template' and method == 'GET':
        return _users_template()
    if method == 'POST':
        return await _import(handler_args, config, _do_users_import)
    return HandlerResult(status=405, body={'error': 'Method Not Allowed'})


async def handle_orgs_import(handler_args: HandlerArgs, kwargs: dict) -> HandlerResult:
    config = kwargs.get('config', {})
    _, err = admin_auth.require_role(handler_args, 'manager')
    if err:
        return err
    parts = _path_parts(handler_args.full_path, _ORG_IMPORT_BASE)
    method = handler_args.method.upper()
    if parts and parts[0] == 'template' and method == 'GET':
        return _orgs_template()
    if method == 'POST':
        return await _import(handler_args, config, _do_orgs_import)
    return HandlerResult(status=405, body={'error': 'Method Not Allowed'})


def _path_parts(full_path: str, base: str):
    path = urlparse(full_path).path
    try:
        rel = PurePath(path).relative_to(PurePath(base))
        return tuple(unquote(p) for p in rel.parts)
    except ValueError:
        return ()


async def _import(handler_args, config, fn):
    body = handler_args.body or {}
    b64 = body.get('file_base64') if isinstance(body, dict) else None
    if not b64:
        return HandlerResult(status=400, body={'error': 'file_base64 필드가 필요합니다'})
    try:
        file_bytes = base64.b64decode(b64)
    except Exception as e:
        return HandlerResult(status=400, body={'error': f'invalid_base64: {e}'})
    base = _csc_base(config)
    token = _bearer(handler_args)
    try:
        result, status = await asyncio.to_thread(fn, file_bytes, token, base)
    except Exception as e:
        logger.log_warning(f"[import] 처리 실패: {e}")
        return HandlerResult(status=500, body={'error': str(e)})
    return HandlerResult(status=status, body=result)


# ── 템플릿 생성 (openpyxl) ──────────────────────────────────────────────
def _users_template() -> HandlerResult:
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'users'
    ws1.append(['name', 'org_code', 'details', 'reject_ids'])
    ws1.append(['홍길동', 'DEV_01', '개발1팀', '+8210001,+8210002'])
    ws2 = wb.create_sheet('volte_subscriptions')
    ws2.append(['name', 'msisdn', 'service_ref', 'imsi', 'password', 'dnd', 'forward_id'])
    ws2.append(['홍길동', '+821357007100', '', '450033100000100', '123456', 'N', ''])
    ws3 = wb.create_sheet('ptt_subscriptions')
    ws3.append(['name', 'msisdn', 'service_ref', 'imsi', 'password', 'dnd'])
    ws3.append(['홍길동', '+82571900100', '', '450033100000100', '123456', 'N'])
    return _xlsx_result(wb, 'cims_import_template.xlsx')


def _orgs_template() -> HandlerResult:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'organizations'
    ws.append(['code', 'name', 'parent_code', 'sort_order'])
    ws.append(['HQ', '본부', '', 1])
    ws.append(['DEV', '개발부', 'HQ', 1])
    ws.append(['DEV_01', '개발1팀', 'DEV', 1])
    return _xlsx_result(wb, 'cims_org_template.xlsx')


def _xlsx_result(wb, filename: str) -> HandlerResult:
    buf = io.BytesIO()
    wb.save(buf)
    return HandlerResult(status=200, body=buf.getvalue(), headers={
        'Content-Type': _XLSX_CT,
        'Content-Disposition': f'attachment; filename="{filename}"',
    })


# ── 오케스트레이션: Excel → csc JSON API ────────────────────────────────
def _do_users_import(file_bytes: bytes, token: str, base: str):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    result = {'created_users': 0, 'created_voip': 0, 'created_ptt': 0, 'errors': []}

    r = _get(base, '/api/v1/users', token)
    if r.status_code != 200:
        return {'error': f'csc 사용자 조회 실패: HTTP {r.status_code}'}, 502
    name_to_id = {u['name']: u['id'] for u in r.json().get('users', [])}

    def _ensure_user(name, row, sheet):
        pid = name_to_id.get(name)
        if pid:
            return pid
        rr = _post(base, '/api/v1/users', token, {'name': name})
        if rr.status_code in (200, 201):
            pid = rr.json().get('id')
            name_to_id[name] = pid
            result['created_users'] += 1
            return pid
        result['errors'].append({'row': row, 'sheet': sheet, 'error': f'사용자 생성 실패 HTTP {rr.status_code}'})
        return None

    if 'users' in wb.sheetnames:
        ws = wb['users']
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            rd = dict(zip(headers, row))
            name = _cell_str(rd.get('name'))
            if not name:
                result['errors'].append({'row': i, 'sheet': 'users', 'error': 'name 필수'})
                continue
            if name in name_to_id:
                continue
            payload = {'name': name, 'org_id': _cell_str(rd.get('org_code')), 'details': _cell_str(rd.get('details'))}
            reject = _cell_str(rd.get('reject_ids'))
            if reject:
                payload['reject_id'] = [x.strip() for x in reject.split(',') if x.strip()]
            rr = _post(base, '/api/v1/users', token, payload)
            if rr.status_code in (200, 201):
                name_to_id[name] = rr.json().get('id')
                result['created_users'] += 1
            else:
                result['errors'].append({'row': i, 'sheet': 'users', 'error': f'HTTP {rr.status_code} {rr.text[:80]}'})

    for sheet, svc, key in (('volte_subscriptions', 'call', 'created_voip'),
                            ('ptt_subscriptions', 'ptt', 'created_ptt')):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            rd = dict(zip(headers, row))
            name = _cell_str(rd.get('name'))
            msisdn = _cell_str(rd.get('msisdn'))
            if not name or not msisdn:
                result['errors'].append({'row': i, 'sheet': sheet, 'error': 'name/msisdn 필수'})
                continue
            pid = _ensure_user(name, i, sheet)
            if not pid:
                continue
            sub = {
                'id': msisdn,
                'imsi': _cell_str(rd.get('imsi')) or msisdn.lstrip('+'),
                'passwd': _cell_str(rd.get('password')) or '123456',
                'dnd': _cell_str(rd.get('dnd')).upper() in ('Y', 'YES', '1', 'TRUE'),
            }
            sref = _cell_str(rd.get('service_ref'))
            if sref:
                sub['service_ref'] = sref
            if svc == 'call':
                sub['forward_id'] = _cell_str(rd.get('forward_id'))
            rr = _post(base, f'/api/v1/users/{pid}/{svc}', token, sub)
            if rr.status_code in (200, 201):
                result[key] += 1
            else:
                result['errors'].append({'row': i, 'sheet': sheet, 'error': f'HTTP {rr.status_code} {rr.text[:80]}'})

    result['total'] = result['created_users'] + result['created_voip'] + result['created_ptt']
    return result, 200


def _do_orgs_import(file_bytes: bytes, token: str, base: str):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    if 'organizations' not in wb.sheetnames:
        return {'error': 'organizations 시트 필요'}, 400
    ws = wb['organizations']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rd = dict(zip(headers, row))
        code = _cell_str(rd.get('code'))
        name = _cell_str(rd.get('name'))
        if not code or not name:
            continue
        try:
            sort_order = int(rd.get('sort_order') or 0)
        except (TypeError, ValueError):
            sort_order = 0
        rows.append({'code': code, 'name': name, 'parent_code': _cell_str(rd.get('parent_code')),
                     'sort_order': sort_order, 'row': i})

    r = _get(base, '/api/v1/organizations', token)
    if r.status_code != 200:
        return {'error': f'csc 조직 조회 실패: HTTP {r.status_code}'}, 502
    code_to_id = {o['code']: o['id'] for o in r.json().get('organizations', [])}

    created, updated, errors = 0, 0, []
    # 1차: code 기준 생성/수정 (parent 는 2차)
    for rd in rows:
        if rd['code'] in code_to_id:
            rr = _put(base, f"/api/v1/organizations/{code_to_id[rd['code']]}", token,
                      {'name': rd['name'], 'sort_order': rd['sort_order']})
            if rr.status_code in (200, 201):
                updated += 1
            else:
                errors.append({'row': rd['row'], 'error': f'HTTP {rr.status_code}'})
        else:
            rr = _post(base, '/api/v1/organizations', token,
                       {'code': rd['code'], 'name': rd['name'], 'sort_order': rd['sort_order']})
            if rr.status_code in (200, 201):
                code_to_id[rd['code']] = rr.json().get('id')
                created += 1
            else:
                errors.append({'row': rd['row'], 'error': f'HTTP {rr.status_code} {rr.text[:80]}'})

    # 2차: parent_id 연결
    for rd in rows:
        if rd['parent_code'] and rd['parent_code'] in code_to_id and rd['code'] in code_to_id:
            _put(base, f"/api/v1/organizations/{code_to_id[rd['code']]}", token,
                 {'parent_id': code_to_id[rd['parent_code']]})

    return {'created': created, 'updated': updated, 'errors': errors}, 200


CIMS_SUBSCRIBER_IMPORT_HANDLER_LIST = [
    (_USERS_IMPORT_BASE, handle_users_import, {}),
    (_ORG_IMPORT_BASE,   handle_orgs_import,  {}),
]
